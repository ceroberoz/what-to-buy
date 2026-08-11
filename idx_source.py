"""IDX official-portal data retrieval for the What-to-Buy valuator.

Fetches the annual financial report index from idx.co.id's GetFinancialReport
API, downloads the XLSX soft-copy attachments, and parses the workbook into
the flat DCF financials schema (Phase 4 D2).

Uses curl_cffi with Chrome impersonation because plain requests gets an HTTP
403 Cloudflare challenge from idx.co.id. The report index reuses the existing
JSON cache helpers from data_source (data/cache/); downloaded XLSX files are
cached under data/cache/idx/.
"""

import json
import os
import urllib.parse

from curl_cffi import requests as cffi_requests

import data_source as ds
from data_source import _cache_get, _cache_set

IDX_BASE = "https://www.idx.co.id"
IDX_API = IDX_BASE + "/primary/ListedCompany/GetFinancialReport"
IDX_CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "cache", "idx")
TIMEOUT = 20
DEFAULT_TTL = 86400


class IdxSourceError(ds.DataSourceError):
    """Raised when IDX data cannot be fetched or downloaded."""


# --- HTTP helpers -----------------------------------------------------------

_session = None


def _get_session():
    """Lazily-created curl_cffi session with Chrome impersonation.

    idx.co.id sits behind Cloudflare and answers plain-requests clients with
    HTTP 403; browser impersonation is what passes the challenge.
    """
    global _session
    if _session is None:
        _session = cffi_requests.Session(impersonate="chrome")
    return _session


# --- IDX report index -------------------------------------------------------

def fetch_report_index(ticker, year, use_cache=True, ttl=DEFAULT_TTL):
    """List of annual report results for a ticker/year from GetFinancialReport.

    Returns the JSON "Results" array; an empty list when the annual report for
    that year has not been filed yet (no "Results" key in the response).
    """
    key = "idx_index_{}_{}".format(ticker.upper(), year)
    if use_cache:
        cached = _cache_get(key, ttl)
        if cached is not None:
            return cached
    params = urllib.parse.urlencode(
        {
            "indexFrom": 0,
            "pageSize": 12,
            "year": year,
            "reportType": "rdf",
            "EmitenType": "s",
            "periode": "audit",
            "kodeEmiten": ticker.upper(),
        }
    )
    url = IDX_API + "?" + params
    response = _get_session().get(url, timeout=TIMEOUT)
    if response.status_code != 200:
        raise IdxSourceError("HTTP {} fetching report index from {}".format(response.status_code, url))
    try:
        payload = json.loads(response.text)
    except (ValueError, json.JSONDecodeError) as exc:
        raise IdxSourceError("invalid JSON from {}: {}".format(url, exc)) from exc
    results = payload.get("Results") or []
    if use_cache:
        _cache_set(key, results)
    return results


# --- XLSX attachment download -----------------------------------------------

def find_xlsx_attachment(result):
    """First attachment dict of a report result whose file type is .xlsx, else None.

    The live API schema uses File_Name / File_Type (and a relative File_Path);
    the soft-copy annual statements are the "FinancialStatement-*-Tahunan-*.xlsx"
    attachments. Skip the XBRL zip bundles.
    """
    for attachment in result.get("Attachments", []):
        file_type = (attachment.get("File_Type") or "").lower()
        file_name = (attachment.get("File_Name") or "").lower()
        if file_type == ".xlsx" or file_name.endswith(".xlsx"):
            return attachment
    return None


def download_xlsx(attachment, dest_dir=None, use_cache=True):
    """Download the XLSX soft copy referenced by an attachment dict to the cache dir.

    Returns the local path. The API File_Path contains raw spaces (and a double
    slash); only the spaces are URL-encoded, the rest of the URL is kept as-is.
    """
    filename = attachment.get("File_Name")
    if not filename:
        raise IdxSourceError("attachment has no File_Name: {}".format(attachment))
    file_url = attachment.get("File_Path")
    if not file_url:
        raise IdxSourceError("attachment {} has no File_Path".format(filename))
    if dest_dir is None:
        dest_dir = IDX_CACHE_DIR
    os.makedirs(dest_dir, exist_ok=True)
    local_path = os.path.join(dest_dir, filename)
    if use_cache and os.path.exists(local_path):
        return local_path
    url = file_url.replace(" ", "%20")
    if url.startswith("/"):
        url = IDX_BASE + url
    response = _get_session().get(url, timeout=TIMEOUT)
    if response.status_code != 200:
        raise IdxSourceError("HTTP {} downloading {} from {}".format(response.status_code, filename, url))
    with open(local_path, "wb") as fh:
        fh.write(response.content)
    return local_path


# --- XLSX soft-copy parser --------------------------------------------------

SHORT_TERM_DEBT_KEYWORDS = (
    "utang bank",
    "jatuh tempo dalam satu tahun",
    "obligasi",
    "sukuk",
    "wesel bayar",
    "surat utang jangka menengah",
    "pinjaman beragunan",
    "pinjaman tanpa agunan",
    "penerusan pinjaman",
    "pinjaman subordinasi",
)
LONG_TERM_DEBT_KEYWORDS = (
    "utang bank",
    "obligasi",
    "sukuk",
    "pinjaman",
    "sewa pembiayaan",
    "wesel bayar",
    "surat utang jangka menengah",
    "utang pembiayaan konsumen",
)
NON_DEBT_LABELS = (
    "utang usaha",
    "utang pajak",
    "utang dividen",
    "utang lainnya",
    "utang nasabah",
    "utang cukai",
    "utang proyek",
    "utang kepada lembaga kliring",
    "utang reasuransi",
    "utang koasuransi",
    "utang pihak berelasi",
    "utang pemegang saham",
    "liabilitas imbalan pasca kerja",
    "pendapatan diterima dimuka",
    "uang jaminan",
    "provisi",
    "liabilitas pajak tangguhan",
    "kontrak liabilitas",
)


def _norm_label(cell_value):
    return str(cell_value).strip().lower()


def _sheet_label_map(ws, label_col=0, value_col=1):
    """{normalized col-A label: value} for the standard label/value sheets.

    Later rows overwrite earlier ones with the same label, so the last
    occurrence of a label wins.
    """
    out = {}
    for row in ws.iter_rows():
        label = row[label_col].value
        if label is not None:
            out[_norm_label(label)] = row[value_col].value
    return out


def _is_debt_row(label, keywords):
    """True when a liability label looks interest-bearing and is not a known non-debt line."""
    if any(skip in label for skip in NON_DEBT_LABELS):
        return False
    return any(keyword in label for keyword in keywords)


def _extract_debt(ws):
    """(short_term, long_term, short_hit, long_hit) interest-bearing debt sums.

    Walks the balance sheet in row order, tracking the current liability
    section, and sums numeric rows whose label matches the debt keywords.
    """
    short_total = 0.0
    long_total = 0.0
    short_hit = False
    long_hit = False
    section = None
    for row in ws.iter_rows():
        label = _norm_label(row[0].value) if row[0].value is not None else ""
        value = row[1].value
        if label == "liabilitas jangka pendek":
            section = "short"
        elif label == "liabilitas jangka panjang":
            section = "long"
        elif label == "jumlah liabilitas jangka panjang":
            section = None
        elif section == "short" and isinstance(value, (int, float)) and _is_debt_row(label, SHORT_TERM_DEBT_KEYWORDS):
            short_total += float(value)
            short_hit = True
        elif section == "long" and isinstance(value, (int, float)) and _is_debt_row(label, LONG_TERM_DEBT_KEYWORDS):
            long_total += float(value)
            long_hit = True
    return short_total, long_total, short_hit, long_hit


def _extract_depreciation(ws):
    """Depreciation additions (col D) from the accumulated-depreciation section of 1611000.

    Returns None when no "Dimiliki langsung" or total "Aset tetap" row with a
    numeric additions column is found inside the section.
    """
    in_accum = False
    fallback = None
    for row in ws.iter_rows():
        tag = _norm_label(row[0].value) if row[0].value is not None else ""
        component = _norm_label(row[1].value) if row[1].value is not None else ""
        additions = row[3].value
        if tag == "akumulasi depresiasi":
            in_accum = True
        elif tag == "aset tetap":
            in_accum = False
        elif in_accum and component == "dimiliki langsung" and isinstance(additions, (int, float)):
            return float(additions)
        elif in_accum and component == "aset tetap" and isinstance(additions, (int, float)):
            fallback = float(additions)
    return fallback


def parse_workbook(path):
    """Parse an IDX annual financial statement XLSX into raw per-sheet values.

    Returns a dict with the raw values in the workbook's stated rounding unit
    (e.g. millions) plus "multiplier", "fiscal_year" and "warnings". Label-
    driven (no hardcoded row numbers). Pure function: no network, no caching.
    """
    import warnings
    warnings.filterwarnings("ignore")
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    parsed = {"warnings": []}

    info = _sheet_label_map(wb["1000000"])
    rounding_text = info.get("pembulatan yang digunakan dalam penyajian jumlah dalam laporan keuangan")
    text = str(rounding_text or "").lower()
    if "jutaan" in text or "million" in text:
        parsed["multiplier"] = 1e6
    elif "ribuan" in text or "thousand" in text:
        parsed["multiplier"] = 1e3
    else:
        parsed["multiplier"] = 1.0
        parsed["warnings"].append("unknown rounding declaration: {}".format(rounding_text))
    parsed["fiscal_year"] = info.get("tanggal akhir periode berjalan")

    income = _sheet_label_map(wb["1321000"])
    parsed["revenue"] = income.get("penjualan dan pendapatan usaha")
    gross_profit = income.get("jumlah laba bruto")
    selling_expenses = income.get("beban penjualan")
    admin_expenses = income.get("beban umum dan administrasi")
    parsed["interest_expense"] = income.get("beban bunga dan keuangan")
    parsed["pre_tax_income"] = income.get("jumlah laba (rugi) sebelum pajak penghasilan")
    tax_expense = income.get("pendapatan (beban) pajak")
    if tax_expense is not None:
        tax_expense = abs(tax_expense)
    parsed["tax_expense"] = tax_expense
    if gross_profit is not None and selling_expenses is not None and admin_expenses is not None:
        parsed["ebit"] = gross_profit - selling_expenses - admin_expenses
    elif parsed["pre_tax_income"] is not None and parsed["interest_expense"] is not None:
        parsed["ebit"] = parsed["pre_tax_income"] + parsed["interest_expense"]
        parsed["warnings"].append("EBIT computed as pre-tax + interest (gross-profit breakdown missing)")
    else:
        parsed["ebit"] = None

    balance = _sheet_label_map(wb["1210000"])
    balance_prev = _sheet_label_map(wb["1210000"], value_col=2)
    parsed["cash"] = balance.get("kas dan setara kas")
    parsed["current_assets"] = balance.get("jumlah aset lancar")
    parsed["current_liabilities"] = balance.get("jumlah liabilitas jangka pendek")
    parsed["current_assets_prev"] = balance_prev.get("jumlah aset lancar")
    parsed["current_liabilities_prev"] = balance_prev.get("jumlah liabilitas jangka pendek")
    parsed["equity"] = balance.get("jumlah ekuitas")
    short_debt, long_debt, short_hit, long_hit = _extract_debt(wb["1210000"])
    parsed["short_term_debt"] = short_debt
    parsed["long_term_debt"] = long_debt
    if not short_hit and not long_hit:
        parsed["warnings"].append("no interest-bearing debt rows matched; assuming 0")

    cashflow = _sheet_label_map(wb["1510000"])
    parsed["operating_cashflow"] = cashflow.get("jumlah arus kas bersih yang diperoleh dari (digunakan untuk) aktivitas operasi")
    capex_parts = (
        cashflow.get("pembayaran untuk perolehan aset tetap"),
        cashflow.get("pembayaran uang muka pembelian aset tetap"),
        cashflow.get("pembayaran untuk perolehan aset takberwujud"),
    )
    if any(part is not None for part in capex_parts):
        parsed["capex"] = sum(part for part in capex_parts if part is not None)
    else:
        parsed["capex"] = None

    depreciation = _extract_depreciation(wb["1611000"])
    parsed["depreciation"] = depreciation
    if depreciation is None:
        parsed["warnings"].append("depreciation movement not found")

    return parsed


def normalize(parsed):
    """Convert parsed workbook values to the flat DCF schema, in full IDR.

    Every numeric field is multiplied by the workbook's rounding multiplier
    (raw values are in millions → full IDR). Shares outstanding are not part
    of the XBRL workbook and stay None (kept from Yahoo in D3).
    """
    multiplier = parsed.get("multiplier", 1.0)
    out = {
        "source": "idx",
        "fiscal_year": parsed.get("fiscal_year"),
        "shares_outstanding": None,
    }
    fields = (
        "revenue",
        "ebit",
        "pre_tax_income",
        "tax_expense",
        "interest_expense",
        "depreciation",
        "capex",
        "operating_cashflow",
        "cash",
        "current_assets",
        "current_liabilities",
        "current_assets_prev",
        "current_liabilities_prev",
        "short_term_debt",
        "long_term_debt",
        "equity",
    )
    for field in fields:
        value = parsed.get(field)
        if value is not None:
            value = float(value) * multiplier
        out[field] = value
    return out
